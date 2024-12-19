from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout, authenticate, login
from .forms import CustomUserCreationForm, CustomAuthenticationForm, ContactForm  # Formularios personalizados definidos en la aplicación.
from django.contrib import messages  # Para manejar mensajes temporales a los usuarios.
from django.urls import reverse  # Para generar URLs a partir de nombres de vistas.
from django.core.mail import EmailMessage  # Para manejar el envío de correos electrónicos.
from django.contrib.auth.models import Group, Permission  # Modelo para manejar grupos y permisos.
from django.contrib.contenttypes.models import ContentType  # Para asignar permisos a los modelos.
from django.core.paginator import Paginator  # Para manejar la paginación de listas.
from django.contrib.auth.decorators import user_passes_test, login_required
from .forms import CustomUserCreationForm, DocenteCreationForm
from .models import Curso, CustomUser
from .forms import CursoForm
from openpyxl import Workbook
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from io import BytesIO
import csv




# Vista para la página de inicio (home)
def index(request):
    return render(request, "core/index.html")  # Renderiza la plantilla "index.html" para la página principal.


# Vista para la página de galería
def galeria(request):
    return render(request, "core/galeria.html")  # Renderiza la plantilla "galeria.html" para la galería de imágenes.


# Vista para la página "Quiénes somos"
def quienes_somos(request):
    return render(request, "core/quienes_somos.html")  # Renderiza la plantilla "quienes_somos.html" para información sobre la organización.



# Vista para la página de servicios
def servicios(request):
    return render(request, "core/servicios.html")  # Renderiza la plantilla "servicios.html" para mostrar los servicios disponibles.




def is_admin(user):
    return user.is_authenticated and user.user_type == 'admin'

def is_docente(user):
    return user.is_authenticated and user.user_type == 'docente'

def is_estudiante(user):
    return user.is_authenticated and user.user_type == 'estudiante'

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})




# Vista para la autenticación de usuarios (inicio de sesión)
def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        print(f"Errores del formulario: {form.errors}")  # Debug temporal
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)  # Mueve esta línea fuera del if
                if user.user_type == 'docente':
                    messages.success(request, f"¡Bienvenido, Profesor {user.username}!")
                    return redirect('ambiente_docente')
                elif user.user_type == 'estudiante':
                    messages.success(request, f"¡Bienvenido, {user.username}!")
                    return redirect('ambiente_estudiante')
                elif user.user_type == 'admin':
                    messages.success(request, f"¡Bienvenido, Admin {user.username}!")
                    return redirect('admin:index')
                else:
                    messages.error(request, "Esta cuenta no tiene un tipo de usuario válido.")
            else:
                messages.error(request, "Usuario o contraseña incorrectos.")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = CustomAuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})

# Vista para cerrar sesión
def logout_view(request):
    logout(request)  # Finaliza la sesión del usuario actual.
    messages.info(request, "Has cerrado sesión exitosamente.")  # Muestra un mensaje informativo.
    return redirect('index')  # Redirige a la página principal.


# Vista para el formulario de contacto
def contacto(request):
    contact_form = ContactForm()  # Crea una instancia vacía del formulario de contacto.

    if request.method == 'POST':  # Si se envía el formulario (método POST).
        contact_form = ContactForm(data=request.POST)  # Instancia el formulario con los datos enviados.

        if contact_form.is_valid():  # Valida el formulario.
            name = request.POST.get('name', '')  # Obtiene el nombre del formulario.
            email = request.POST.get('email', '')  # Obtiene el correo del formulario.
            message = request.POST.get('message', '')  # Obtiene el mensaje del formulario.

            # Configura el correo electrónico.
            email = EmailMessage(
                'Mensaje de contacto recibido',
                'Mensaje enviado por {} <{}>:\n\n{}'.format(name, email, message),
                email,
                ['800e954f00b339@inbox.mailtrap.io'],  # Dirección de destino para pruebas.
                reply_to=[email],
            )

            try:
                email.send()  # Envía el correo electrónico.
                return redirect(reverse('contacto') + '?ok')  # Redirige con un mensaje de éxito si el envío es exitoso.
            except:
                return redirect(reverse('contacto') + '?error')  # Redirige con un mensaje de error si el envío falla.

    return render(request, "core/contacto.html", {'form': contact_form})  # Renderiza la plantilla de contacto con el formulario.

@user_passes_test(is_admin)
def crear_docente(request):
    if request.method == 'POST':
        form = DocenteCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_docentes')
    else:
        form = DocenteCreationForm()
    return render(request, 'crear_docente.html', {'form': form})


@login_required
def lista_cursos(request):
    cursos = Curso.objects.all()
    return render(request, 'cursos/lista_cursos.html', {'cursos': cursos})

@login_required
@user_passes_test(lambda u: u.user_type == 'estudiante')
def mis_cursos(request):
    cursos = request.user.cursos_inscritos.all()
    return render(request, 'cursos/mis_cursos.html', {'cursos': cursos})

@login_required
@user_passes_test(lambda u: u.user_type == 'docente')
def gestionar_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id, docente=request.user)
    if request.method == 'POST':
        curso.nombre = request.POST.get('nombre')
        curso.descripcion = request.POST.get('descripcion')
        curso.save()
        messages.success(request, 'Curso actualizado exitosamente')
        return redirect('gestionar_curso', curso_id=curso.id)
    return render(request, 'cursos/gestionar_curso.html', {'curso': curso})



@login_required
def ambiente(request):
    """Vista que redirige según el tipo de usuario"""
    user_type = request.user.user_type
    if user_type == 'estudiante':
        return redirect('ambiente_estudiante')
    elif user_type == 'docente':
        return redirect('ambiente_docente')
    elif user_type == 'admin':
        return redirect('ambiente_admin')
    else:
        messages.error(request, "Tipo de usuario no válido")
        return redirect('index')

@login_required
def ambiente_estudiante(request):
    """Vista del ambiente del estudiante"""
    cursos_inscritos = request.user.cursos_inscritos.all()
    cursos_disponibles = Curso.objects.exclude(estudiantes=request.user)
    
    context = {
        'cursos_inscritos': cursos_inscritos,
        'cursos_disponibles': cursos_disponibles,
    }
    return render(request, 'core/ambiente_estudiante.html', context)

@login_required
def detalle_curso(request, curso_id):
    """Vista para ver detalles de un curso"""
    curso = get_object_or_404(Curso, id=curso_id)
    context = {
        'curso': curso,
    }
    return render(request, 'core/detalle_curso.html', context)

@login_required
def inscribir_curso(request, curso_id):
    """Vista para inscribirse a un curso"""
    if request.user.user_type != 'estudiante':
        messages.error(request, "Solo los estudiantes pueden inscribirse a cursos")
        return redirect('ambiente')
        
    curso = get_object_or_404(Curso, id=curso_id)
    curso.estudiantes.add(request.user)
    messages.success(request, f"Te has inscrito exitosamente en {curso.nombre}")
    return redirect('ambiente_estudiante')

@login_required
@user_passes_test(lambda u: u.user_type == 'docente')
def ambiente_docente(request):
    """Vista del ambiente del docente"""
    # Obtener solo los cursos donde el docente actual es el profesor
    cursos = Curso.objects.filter(docente=request.user)
    
    # Agregar estadísticas básicas
    for curso in cursos:
        curso.total_estudiantes = curso.estudiantes.count()
    
    context = {
        'cursos': cursos,
    }
    return render(request, 'core/ambiente_docente.html', context)

@login_required
@user_passes_test(lambda u: u.user_type == 'docente')
def editar_curso(request, curso_id):
    """Vista para editar un curso"""
    curso = get_object_or_404(Curso, id=curso_id, docente=request.user)
    
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        
        if nombre and descripcion:
            curso.nombre = nombre
            curso.descripcion = descripcion
            curso.save()
            messages.success(request, 'Curso actualizado exitosamente')
            return redirect('ambiente_docente')
        else:
            messages.error(request, 'Por favor completa todos los campos')
            
    context = {
        'curso': curso
    }
    return render(request, 'core/editar_curso.html', context)

@login_required
@user_passes_test(lambda u: u.user_type == 'docente')
def lista_alumnos_excel(request, curso_id):
    """Exportar lista de alumnos a Excel"""
    curso = get_object_or_404(Curso, id=curso_id, docente=request.user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Alumnos"
    
    # Encabezados
    headers = ['Nombre', 'Apellido', 'Email', 'Username']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos de alumnos
    for row, estudiante in enumerate(curso.estudiantes.all(), 2):
        ws.cell(row=row, column=1, value=estudiante.first_name)
        ws.cell(row=row, column=2, value=estudiante.last_name)
        ws.cell(row=row, column=3, value=estudiante.email)
        ws.cell(row=row, column=4, value=estudiante.username)
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="lista_alumnos_{curso.nombre}.xlsx"'
    wb.save(response)
    return response

@login_required
@user_passes_test(lambda u: u.user_type == 'docente')
def lista_alumnos_pdf(request, curso_id):
    """Exportar lista de alumnos a PDF"""
    curso = get_object_or_404(Curso, id=curso_id, docente=request.user)
    
    # Crear buffer de memoria
    buffer = BytesIO()
    
    # Crear el PDF
    p = canvas.Canvas(buffer, pagesize=letter)
    p.setTitle(f"Lista de Alumnos - {curso.nombre}")
    
    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 750, f"Lista de Alumnos - {curso.nombre}")
    
    # Información del curso
    p.setFont("Helvetica", 12)
    p.drawString(50, 720, f"Total de estudiantes: {curso.estudiantes.count()}")
    
    # Encabezados
    y = 680
    p.line(50, y, 550, y)
    y -= 20
    p.drawString(50, y, "Nombre")
    p.drawString(200, y, "Apellido")
    p.drawString(350, y, "Email")
    y -= 10
    p.line(50, y, 550, y)
    
    # Lista de estudiantes
    y -= 20
    p.setFont("Helvetica", 10)
    for estudiante in curso.estudiantes.all():
        if y < 50:  # Nueva página si no hay espacio
            p.showPage()
            p.setFont("Helvetica", 10)
            y = 750
            
        p.drawString(50, y, estudiante.first_name)
        p.drawString(200, y, estudiante.last_name)
        p.drawString(350, y, estudiante.email)
        y -= 20
    
    p.showPage()
    p.save()
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lista_alumnos_{curso.nombre}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response